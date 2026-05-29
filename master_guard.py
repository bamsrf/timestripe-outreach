"""
Shared safeguards for MASTER_youtube_outreach.xlsx writes:

1. backup_master(reason) — copies MASTER to output/master_backups/ with timestamp
   + reason tag. Call before ANY write that could lose data.

2. sanity_check_status(expected_minimum) — counts non-empty Status rows in MASTER.
   If much fewer than expected (e.g. half lost), raise to abort the operation.

Usage:
    from master_guard import backup_master, sanity_check_status
    backup_master("before-send")
    sanity_check_status(expected_minimum=30)  # raises if catastrophic loss
"""
import csv
import shutil
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).parent
MASTER = ROOT / "output" / "MASTER_youtube_outreach.xlsx"
BACKUP_DIR = ROOT / "output" / "master_backups"
LOG_FILE = ROOT / "output" / "outreach_log.csv"


def backup_master(reason: str = "auto") -> Path:
    """Copy MASTER xlsx to backup dir with timestamp + reason tag."""
    if not MASTER.exists():
        return None
    BACKUP_DIR.mkdir(exist_ok=True)
    safe_reason = "".join(c if c.isalnum() or c in "-_" else "_" for c in reason)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = BACKUP_DIR / f"MASTER_{ts}_{safe_reason}.xlsx"
    shutil.copy(MASTER, dest)
    # Prune: keep last 30 backups
    backups = sorted(BACKUP_DIR.glob("MASTER_*.xlsx"))
    for old in backups[:-30]:
        try: old.unlink()
        except OSError: pass
    return dest


def _expected_sent_count() -> int:
    """How many distinct emails should have Status non-empty based on send log?"""
    if not LOG_FILE.exists():
        return 0
    seen = set()
    with open(LOG_FILE) as f:
        for r in csv.DictReader(f):
            if r.get("outcome") == "sent":
                em = r.get("email", "").lower().strip()
                if em and "@srv1.mail-tester.com" not in em and "vladisru" not in em:
                    seen.add(em)
    return len(seen)


def sanity_check_status(tolerance_pct: float = 30.0) -> tuple[bool, str]:
    """Count rows in MASTER with non-empty Status. Compare to expected from log.
    Return (is_ok, message). If is_ok=False, caller should ABORT and warn."""
    from openpyxl import load_workbook
    if not MASTER.exists():
        return True, "no MASTER yet"
    wb = load_workbook(MASTER, read_only=True)
    ws = wb.active
    hs = [c.value for c in ws[1]]
    if "Status" not in hs:
        return True, "no Status column"
    st_i = hs.index("Status")
    actual = sum(1 for row in ws.iter_rows(min_row=2, values_only=True)
                 if row[st_i] and str(row[st_i]).strip())
    expected = _expected_sent_count()
    if expected == 0:
        return True, "log is empty (first ever run)"

    deficit = expected - actual
    deficit_pct = (deficit / expected * 100) if expected else 0

    if deficit_pct > tolerance_pct:
        return False, (
            f"⚠️ STATUS SANITY FAIL: log says we sent to {expected} contacts, "
            f"but MASTER has only {actual} with non-empty Status. "
            f"Missing {deficit} ({deficit_pct:.0f}%). "
            f"Looks like MASTER was reset since last send. "
            f"ABORTING to prevent duplicate sends. "
            f"Run repair_master.py or restore from output/master_backups/."
        )
    return True, f"ok: log={expected}, MASTER={actual} ({deficit} diff, within tolerance)"
