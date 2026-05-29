#!/usr/bin/env python3
"""
Mailbox sync — pulls in replies AND outbound manual sends from Yandex.

Two passes:
  - INBOX scan: senders match a MASTER row in {Sent intro, Sent program info}
                → set Status="Replied"
  - SENT scan: recipients match a MASTER row with empty Status
                → set Status="Sent intro", First Contact Date=message date

Usage:
    python reply_detector.py                # dry-run (show what would change)
    python reply_detector.py --apply        # actually update MASTER
    python reply_detector.py --since 30     # look back 30 days
    python reply_detector.py --no-outbound  # only check INBOX, skip Sent
    python reply_detector.py --no-inbound   # only check Sent, skip INBOX
"""
import argparse
import email
import imaplib
import os
import re
import sys
from datetime import datetime, timedelta, date as _date
from email.utils import getaddresses, parsedate_to_datetime
from pathlib import Path

from dotenv import load_dotenv

ROOT = Path(__file__).parent
load_dotenv(ROOT / ".env")

REQUIRED = ["SMTP_USER", "SMTP_PASSWORD"]
for k in REQUIRED:
    if not os.environ.get(k):
        print(f"ERROR: {k} missing in .env", file=sys.stderr)
        sys.exit(1)

# Yandex folder names (UTF-7 IMAP)
YANDEX_INBOX = "INBOX"
YANDEX_SENT_CANDIDATES = ["Sent", "Отправленные", "&BB4EQgQ,BEAEMAQyBDsENQQ9BD0ESwQ1-"]  # last = IMAP UTF-7


def _open_imap():
    host = "imap.yandex.ru"
    user = os.environ["SMTP_USER"]
    pwd = os.environ["SMTP_PASSWORD"]
    imap = imaplib.IMAP4_SSL(host, 993, timeout=30)
    imap.login(user, pwd)
    return imap, user


def _find_sent_folder(imap):
    """Yandex names Sent folder differently per language. Try candidates."""
    for name in YANDEX_SENT_CANDIDATES:
        ok, _ = imap.select(name, readonly=True)
        if ok == "OK":
            return name
    # Fallback: list folders, find one with \\Sent flag
    ok, folders = imap.list()
    if ok == "OK":
        for f in folders:
            if b"\\Sent" in f:
                # parse folder name out of "(...) "DELIM" "NAME"" response
                m = re.search(rb'"([^"]+)"$', f)
                if m:
                    return m.group(1).decode("utf-8", errors="replace")
    return None


def _scan_folder_for_addresses(imap, folder, days_back, header_field, skip_replies=False):
    """Search messages in folder since N days back, return list of
    (address_lowercase, message_date) tuples. header_field: 'From' or 'To'.

    If skip_replies=True, messages whose Subject starts with 'Re:' / 'Fwd:'
    are excluded — they are user-replies to bloggers, NOT new outreach,
    so they should not be misinterpreted as fresh sends. (Fixes 24.05 bug
    where user-replies overwrote 'Replied' status back to 'Sent intro'.)"""
    ok, _ = imap.select(folder, readonly=True)
    if ok != "OK":
        print(f"! could not open folder {folder!r}", file=sys.stderr)
        return []
    since = (datetime.now() - timedelta(days=days_back)).strftime("%d-%b-%Y")
    ok, ids = imap.search(None, f'(SINCE {since})')
    if ok != "OK":
        return []
    ids = ids[0].split() if ids and ids[0] else []
    print(f"  Folder {folder!r}: {len(ids)} messages since {since}")

    skipped_replies = 0
    found = []
    for i in range(0, len(ids), 100):
        batch = b",".join(ids[i:i+100])
        # Include Subject for reply detection
        ok, data = imap.fetch(batch, f"(BODY[HEADER.FIELDS ({header_field} DATE SUBJECT)])")
        if ok != "OK":
            continue
        for item in data:
            if not isinstance(item, tuple):
                continue
            hdrs = email.message_from_bytes(item[1])
            raw = hdrs.get(header_field, "")
            subj = (hdrs.get("Subject", "") or "").strip()
            if skip_replies and re.match(r"^(re|fwd?|fw)\s*:", subj, re.IGNORECASE):
                skipped_replies += 1
                continue
            msg_date = None
            try:
                if hdrs.get("Date"):
                    msg_date = parsedate_to_datetime(hdrs["Date"]).date()
            except Exception:
                pass
            if not msg_date:
                msg_date = datetime.now().date()
            for _, addr in getaddresses([raw]):
                if addr and "@" in addr:
                    found.append((addr.lower().strip(), msg_date))
    if skipped_replies:
        print(f"  (skipped {skipped_replies} messages with Re:/Fwd: subject — user-replies, not outreach)")
    return found


def _load_master(master_path):
    from openpyxl import load_workbook
    wb = load_workbook(master_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    return wb, ws, headers


def _col(headers, name):
    try:
        return headers.index(name)
    except ValueError:
        raise SystemExit(f"ERROR: '{name}' column not found in master")


def apply_changes(master_path, replied_addrs, sent_to_addrs, apply):
    wb, ws, headers = _load_master(master_path)
    email_i = _col(headers, "Email")
    status_i = _col(headers, "Status")
    fcd_i = _col(headers, "First Contact Date")
    lastfu_i = _col(headers, "Last Follow-up")

    # Build lookup: email → latest message date
    replied_dates = {}
    for addr, d in replied_addrs:
        if addr not in replied_dates or d > replied_dates[addr]:
            replied_dates[addr] = d
    sent_dates = {}
    for addr, d in sent_to_addrs:
        if addr not in sent_dates or d > sent_dates[addr]:
            sent_dates[addr] = d

    today = datetime.now().date()
    repl_changes, sent_changes = [], []

    for row in ws.iter_rows(min_row=2):
        em = (row[email_i].value or "").lower().strip()
        if not em:
            continue
        status = (row[status_i].value or "").strip()
        name = row[headers.index("Name")].value

        # Pass 1: replies
        if em in replied_dates and status == "Sent intro":
            repl_changes.append((row, name, em, status))

        # Pass 2: outbound (manual send detected in Sent folder)
        if em in sent_dates and status in ("", "Not contacted"):
            sent_changes.append((row, name, em, sent_dates[em]))

    print(f"\n=== Pass 1: replies → {len(repl_changes)} contacts to flag as 'Replied' ===")
    for row, name, em, prev in repl_changes:
        print(f"  {name[:35]:<35} <{em}>  was: {prev:<20} → Replied")
        if apply:
            row[status_i].value = "Replied"
            row[lastfu_i].value = today

    print(f"\n=== Pass 2: outbound → {len(sent_changes)} contacts manually contacted ===")
    for row, name, em, d in sent_changes:
        print(f"  {name[:35]:<35} <{em}>  (sent {d})  → Sent intro")
        if apply:
            row[status_i].value = "Sent intro"
            row[fcd_i].value = d
            row[lastfu_i].value = d

    if apply and (repl_changes or sent_changes):
        # Backup before writing
        try:
            from master_guard import backup_master
            backup_master(reason="reply-detector")
        except ImportError:
            pass
        # Make sure date columns have proper number format
        from openpyxl.utils import get_column_letter
        for ci_name in ("First Contact Date", "Last Follow-up"):
            ci = headers.index(ci_name) + 1
            for cell in ws[get_column_letter(ci)][1:]:
                cell.number_format = "yyyy-mm-dd"
        wb.save(master_path)
        print(f"\n✓ Updated {master_path.name}")
        # Push changes to Google Sheet + repaint colors
        import subprocess as _sp, sys as _sys
        try:
            _sp.run([_sys.executable, str(ROOT / "sheet_status_sync.py"), "--paint"],
                    cwd=str(ROOT), check=False, timeout=120)
            print("✓ Synced to Google Sheet")
        except Exception as e:
            print(f"! Sheet sync failed: {e} (master is still up to date)")
    elif repl_changes or sent_changes:
        print("\n(dry-run — pass --apply to actually update MASTER)")
    else:
        print("Nothing to update.")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--since", type=int, default=14, help="days to look back")
    parser.add_argument("--apply", action="store_true", help="actually update MASTER")
    parser.add_argument("--master", default="output/MASTER_youtube_outreach.xlsx")
    parser.add_argument("--no-inbound", action="store_true", help="skip INBOX scan")
    parser.add_argument("--no-outbound", action="store_true", help="skip Sent folder scan")
    args = parser.parse_args()

    master = ROOT / args.master if not Path(args.master).is_absolute() else Path(args.master)
    if not master.exists():
        print(f"ERROR: master file not found at {master}", file=sys.stderr)
        sys.exit(1)

    imap, user = _open_imap()
    print(f"Connected as {user}")
    try:
        replied = []
        sent_to = []
        if not args.no_inbound:
            print("\nScanning INBOX for replies...")
            replied = _scan_folder_for_addresses(imap, YANDEX_INBOX, args.since, "From")
            # Exclude messages from yourself (auto-replies, threads with self)
            replied = [(a, d) for a, d in replied if a != user.lower()]
            print(f"  Unique reply senders: {len({a for a,_ in replied})}")
        if not args.no_outbound:
            print("\nScanning Sent folder for outbound...")
            sent_folder = _find_sent_folder(imap)
            if sent_folder:
                print(f"  Detected Sent folder: {sent_folder!r}")
                # skip_replies=True: ignore Re:/Fwd: messages — those are user-replies
                # to bloggers, not new outreach (would otherwise corrupt Status).
                sent_to = _scan_folder_for_addresses(imap, sent_folder, args.since, "To",
                                                    skip_replies=True)
                # Exclude messages sent to yourself / sent from outreach_sender.py
                # (the script also IMAP-appends to "Bloggers", not Sent, so this is safe)
                sent_to = [(a, d) for a, d in sent_to if a != user.lower()]
                print(f"  Unique sent-to recipients: {len({a for a,_ in sent_to})}")
            else:
                print("  ! could not find Sent folder")
    finally:
        try:
            imap.logout()
        except Exception:
            pass

    apply_changes(master, replied, sent_to, apply=args.apply)


if __name__ == "__main__":
    main()
