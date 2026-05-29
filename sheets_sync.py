#!/usr/bin/env python3
"""
Google Sheets sync — bidirectional sync between local MASTER xlsx and your
Google Sheet (tab "Vlad").

Two modes:
  pull   — read existing emails from sheet, save to crm_emails.csv
           (used by outreach_sender.py to skip already-CRM contacts)
  push   — append NEW contacts from MASTER xlsx to Google Sheet (default mode)
           Skips contacts whose email is already in the sheet.
           New rows go AFTER the last filled row — existing data is never touched.

Setup (one-time):
    1. Google Cloud: enable "Google Sheets API"
    2. Create Service Account, generate JSON key, save to ~/.config/timestripe-gsheets.json
    3. Share your Google Sheet with the service account email (Editor role)
    4. Add GSHEETS_ID=<your sheet id> to .env

Usage:
    python sheets_sync.py                       # push new contacts (default)
    python sheets_sync.py --dry-run             # show what would be added
    python sheets_sync.py --mode pull           # just refresh crm_emails.csv
    python sheets_sync.py --tab "Vlad"          # specify tab
    python sheets_sync.py --source path/to.xlsx # different source
"""
import argparse
import csv
import os
import sys
from datetime import date, datetime
from pathlib import Path

import gspread
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook

ROOT = Path(__file__).parent
load_dotenv(ROOT / ".env")

DEFAULT_CREDS_PATH = Path.home() / ".config" / "timestripe-gsheets.json"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]  # read+write


def open_sheet():
    sheet_id = os.environ.get("GSHEETS_ID")
    if not sheet_id:
        sys.exit("ERROR: set GSHEETS_ID in .env (your Google Sheet ID)")
    creds_path = Path(os.environ.get("GSHEETS_CREDS", DEFAULT_CREDS_PATH))
    if not creds_path.exists():
        sys.exit(f"ERROR: credentials JSON not found at {creds_path}")
    creds = Credentials.from_service_account_file(str(creds_path), scopes=SCOPES)
    gc = gspread.authorize(creds)
    return gc.open_by_key(sheet_id)


def find_header_row(ws):
    """Scan first 10 rows looking for a row whose first cell is 'Name' — that's
    the column-header row in this sheet."""
    rows = ws.get_all_values()[:10]
    for i, row in enumerate(rows, 1):
        if row and row[0].strip().lower() == "name":
            return i, [h.strip() for h in row if h.strip()]
    sys.exit("ERROR: could not find a header row starting with 'Name' in first 10 rows of the tab")


def get_existing_emails(ws, header_row_num, email_col_idx):
    """Read all email values from rows below the header row."""
    all_rows = ws.get_all_values()
    emails = set()
    for row in all_rows[header_row_num:]:
        if len(row) > email_col_idx and row[email_col_idx].strip():
            emails.add(row[email_col_idx].strip().lower())
    return emails


def stringify(value):
    """Coerce any value to a string suitable for Google Sheets cell."""
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value)


def load_master_rows(source_path):
    """Load rows from MASTER xlsx as list of dicts keyed by header."""
    wb = load_workbook(source_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, r)))
    return rows


def push_mode(args):
    sh = open_sheet()
    try:
        ws = sh.worksheet(args.tab)
    except gspread.WorksheetNotFound:
        sys.exit(f"ERROR: tab {args.tab!r} not found in spreadsheet")

    header_row_num, sheet_headers = find_header_row(ws)
    print(f"Sheet: {sh.title!r} / tab {args.tab!r}")
    print(f"Headers found at row {header_row_num}: {sheet_headers}")

    if "Email" not in sheet_headers:
        sys.exit("ERROR: sheet has no 'Email' column")
    email_col_idx = sheet_headers.index("Email")

    existing_emails = get_existing_emails(ws, header_row_num, email_col_idx)
    last_data_row = header_row_num + sum(1 for _ in ws.get_all_values()[header_row_num:])
    print(f"Existing contacts in sheet: {len(existing_emails)}")
    print(f"Next empty row: {last_data_row + 1}")

    source = ROOT / args.source if not Path(args.source).is_absolute() else Path(args.source)
    if not source.exists():
        sys.exit(f"ERROR: source not found at {source}")
    master_rows = load_master_rows(source)
    print(f"MASTER rows: {len(master_rows)}")

    # Filter: only rows with email NOT yet in sheet
    to_add = []
    for r in master_rows:
        em = (r.get("Email") or "").strip().lower()
        if not em or em in existing_emails:
            continue
        # Build row in sheet's column order
        row_values = [stringify(r.get(h, "")) for h in sheet_headers]
        to_add.append(row_values)

    print(f"New contacts to append: {len(to_add)}")

    if not to_add:
        print("Nothing to add — sheet already has all MASTER contacts.")
        return

    # Show preview
    print("\nPreview (first 5):")
    for row in to_add[:5]:
        print(f"  {row[0][:30]:<30} | {row[email_col_idx]:<35} | {row[sheet_headers.index('Platform')] if 'Platform' in sheet_headers else '-':<8} | Tier {row[sheet_headers.index('Tier')] if 'Tier' in sheet_headers else '-'}")

    if args.dry_run:
        print("\n(dry-run — nothing written)")
        return

    # Append in chunks (gspread handles batching internally)
    BATCH = 500
    for i in range(0, len(to_add), BATCH):
        chunk = to_add[i:i + BATCH]
        ws.append_rows(chunk, value_input_option="USER_ENTERED")
        print(f"  ✓ appended {len(chunk)} rows")

    print(f"\n✓ Pushed {len(to_add)} new contacts to {sh.title!r}/{args.tab!r}")

    # Also refresh crm_emails.csv so outreach_sender doesn't try to re-send
    pull_mode(args, _sh=sh, _ws=ws, _header_row=header_row_num, _email_col=email_col_idx)


def pull_mode(args, _sh=None, _ws=None, _header_row=None, _email_col=None):
    if _sh is None:
        sh = open_sheet()
        ws = sh.worksheet(args.tab)
        header_row_num, sheet_headers = find_header_row(ws)
        if "Email" not in sheet_headers:
            sys.exit("ERROR: sheet has no 'Email' column")
        email_col_idx = sheet_headers.index("Email")
    else:
        sh, ws, header_row_num, email_col_idx = _sh, _ws, _header_row, _email_col

    sheet_emails = get_existing_emails(ws, header_row_num, email_col_idx)

    # IMPORTANT: crm_emails.csv represents contacts in user's CRM that are NOT
    # in MASTER xlsx (i.e. manually added). We exclude those from outreach.
    # Contacts that ARE in MASTER get their state tracked via Status column,
    # so they should NOT be in this exclusion list.
    source = ROOT / args.source if not Path(args.source).is_absolute() else Path(args.source)
    master_emails = set()
    if source.exists():
        from openpyxl import load_workbook
        wb = load_workbook(source)
        ws_m = wb.active
        m_headers = [c.value for c in ws_m[1]]
        m_email_i = m_headers.index("Email") if "Email" in m_headers else None
        if m_email_i is not None:
            for row in ws_m.iter_rows(min_row=2, values_only=True):
                em = (row[m_email_i] or "")
                if em:
                    master_emails.add(str(em).lower().strip())

    manual_only = sheet_emails - master_emails

    out = ROOT / args.out
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["email"])
        w.writeheader()
        for e in sorted(manual_only):
            w.writerow({"email": e})
    print(f"→ Wrote {len(manual_only)} manual-only emails to {out.name} "
          f"(sheet has {len(sheet_emails)}, MASTER has {len(master_emails)})")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["push", "pull"], default="push",
                        help="push (default): add new MASTER rows to sheet; pull: just refresh crm_emails.csv")
    parser.add_argument("--tab", default="Vlad", help="sheet tab name")
    parser.add_argument("--source", default="output/MASTER_youtube_outreach.xlsx",
                        help="MASTER xlsx to read new contacts from")
    parser.add_argument("--out", default="output/crm_emails.csv",
                        help="local file with existing CRM emails (used by outreach_sender)")
    parser.add_argument("--dry-run", action="store_true", help="show what would happen")
    args = parser.parse_args()

    if args.mode == "push":
        push_mode(args)
    else:
        pull_mode(args)


if __name__ == "__main__":
    main()
