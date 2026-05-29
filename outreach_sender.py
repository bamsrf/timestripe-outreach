#!/usr/bin/env python3
"""
TimeStripe outreach sender.

Reads MASTER_youtube_outreach.xlsx (or any CSV/XLSX with CRM columns),
filters by fit_score and Status, renders per-tier template, sends via SMTP
(Gmail / Yandex / Yandex 360 — any standard SMTP).

DEFAULT MODE IS DRY-RUN. Nothing is sent until you pass --send.

Usage:
  python outreach_sender.py                          # dry-run, shows what would be sent
  python outreach_sender.py --send                   # actually sends
  python outreach_sender.py --send --limit 5         # send max 5 (for testing)
  python outreach_sender.py --send --tier A          # only Tier A
  python outreach_sender.py --source output/youtube_20260520_1536.csv

Status tracking:
  After each successful send, the source file gets Status="Sent intro" +
  First Contact Date=today written back. A persistent log is also appended
  to output/outreach_log.csv.
"""
import argparse
import csv
import imaplib
import os
import random
import re
import smtplib
import ssl
import sys
import time
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr, formatdate, make_msgid
from pathlib import Path

import yaml
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv(Path(__file__).parent / ".env")

# Common English first names — used to detect names inside compound usernames
# like "Queennjasmine" → "Jasmine".
COMMON_FIRST_NAMES = {
    "alex","alice","amelia","amy","ana","anna","anya","asma","ashley","ayu",
    "ben","beth","bina","brad","brett","brian","caitlin","cara","carlie",
    "charlotte","chloe","chris","claire","clara","claudia","cole","conor",
    "dan","dave","david","dean","ella","ellie","emily","emma","erica","erin",
    "ethan","eva","fiona","fred","gabby","gabe","gina","grace","hailey",
    "hannah","haruko","heidi","helen","ian","irene","isa","isabel","jack",
    "jake","james","jane","jason","jasmine","jenny","jess","jessica","jill",
    "john","jules","julia","kate","kayla","kelly","kim","kira","kyle","lara",
    "laura","lavender","leah","liam","lily","lisa","lucia","lucy","luke",
    "macy","maddie","mali","mara","mark","martha","matt","max","maya","megan",
    "melia","mia","mike","miki","miles","mira","mona","nao","nate","nick",
    "nicole","nina","noah","olivia","owen","paige","paul","pete","peter",
    "prity","quinn","rachel","rebecca","rita","ryan","sam","sara","sarah",
    "shanna","sofia","sofi","sophia","soso","sree","stephanie","summer","tara",
    "tess","tom","tristan","umaya","vlad","wafa","will","wendy","yvonne","zoe",
    "anya","navi","navami","jacob","tanzila","akumi","eshaan","mindro",
}

# Words to drop when looking at channel name tokens
STOP_TOKENS = {"the","my","with","by","and","on","at","of","is","this","that",
               "life","planner","planning","journey","studio","art","craft",
               "world","creator","tv","yt","official","channel","vlogs","blog",
               "youtube","tiktok","day","daily","online","real","live"}


def smart_extract_name(channel_name, email):
    """Pull a likely first name from channel name or email local-part.
    Returns the name (title-cased) or empty string if nothing reasonable found."""
    cn = (channel_name or "").strip()

    # 1) "X's Diary" / "Erin's notes" → "X"
    m = re.search(r"\b([A-Za-z]{3,})['’]s\b", cn)
    if m:
        return m.group(1).capitalize()

    # 2) "Something with NAME" / "by NAME" / "& NAME"
    m = re.search(r"\b(?:with|by|&|and|feat)\s+([A-Za-z]{3,})\b", cn, re.IGNORECASE)
    if m:
        cand = m.group(1)
        if cand.lower() not in STOP_TOKENS:
            return cand.capitalize()

    # 3) Tokenize channel name (split on spaces, dots, underscores, dashes)
    tokens = [t for t in re.split(r"[\s._\-|/]+", cn) if t]
    # 3a) Split camelCase tokens: "JashiiCorrin" → ["Jashii","Corrin"]
    split_tokens = []
    for t in tokens:
        parts = re.findall(r"[A-Z][a-z]+|[a-z]+|[A-Z]+(?=[A-Z][a-z]|$)", t) or [t]
        split_tokens.extend(parts)
    cleaned = [t for t in split_tokens if t.isalpha() and t.lower() not in STOP_TOKENS]

    # 4) Pick token that's a known first name
    for t in cleaned:
        if t.lower() in COMMON_FIRST_NAMES:
            return t.capitalize()

    # 5) If only one alpha token left after stop-word removal — use it
    if len(cleaned) == 1 and 3 <= len(cleaned[0]) <= 15:
        return cleaned[0].capitalize()

    # 6) Email local-part — search inside for a known first name
    local = (email or "").split("@")[0].lower()
    for name in COMMON_FIRST_NAMES:
        if name in local and len(name) >= 4:
            return name.capitalize()

    # 7) Fallback: empty string (caller decides)
    return ""

REQUIRED_ENV = ["SMTP_HOST", "SMTP_PORT", "SMTP_USER", "SMTP_PASSWORD"]
for k in REQUIRED_ENV:
    if not os.environ.get(k):
        print(f"ERROR: {k} missing in .env", file=sys.stderr)
        sys.exit(1)

CRM_COLUMNS = [
    "Name", "Email", "Platform", "URL", "Followers", "Tier", "Niche",
    "Status", "First Contact Date", "Last Follow-up", "Notes",
    "Program Type", "Videos Published", "Total Views",
    "Payout Owed", "Payout Sent", "ROI",
    "fit_score", "fit_flags",
]


def load_rows(path):
    """Load rows from CSV or XLSX. Return (rows, headers, format)."""
    path = Path(path)
    if path.suffix.lower() == ".csv":
        with open(path, encoding="utf-8") as f:
            rdr = csv.DictReader(f)
            headers = rdr.fieldnames
            rows = list(rdr)
        return rows, headers, "csv"
    elif path.suffix.lower() == ".xlsx":
        wb = load_workbook(path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(r[i]) if r[i] is not None else "") for i in range(len(headers))})
        return rows, headers, "xlsx"
    else:
        raise ValueError(f"Unsupported source format: {path.suffix}")


DATE_COLUMNS = {"First Contact Date", "Last Follow-up"}


def _coerce_date(value):
    """Convert 'YYYY-MM-DD' string to datetime.date so Excel treats it as a real date."""
    from datetime import date as _date
    if isinstance(value, _date):
        return value
    if not value or not str(value).strip():
        return ""
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except ValueError:
        return value  # leave as-is if not parseable


def save_rows(path, rows, headers, fmt):
    """Save back to original format, preserving styling for xlsx."""
    if fmt == "csv":
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
            w.writeheader()
            w.writerows(rows)
    else:
        wb = Workbook(); ws = wb.active; ws.title = "Outreach"
        ws.append(headers)
        for r in rows:
            row_out = []
            for h in headers:
                v = r.get(h, "")
                if h in DATE_COLUMNS:
                    v = _coerce_date(v)
                row_out.append(v)
            ws.append(row_out)
        hf = Font(bold=True, color="FFFFFF"); hb = PatternFill("solid", fgColor="4F81BD")
        for cell in ws[1]:
            cell.font = hf; cell.fill = hb; cell.alignment = Alignment(horizontal="center")
        # Apply date format to date cells
        for ci, h in enumerate(headers, 1):
            if h in DATE_COLUMNS:
                for cell in ws[get_column_letter(ci)][1:]:
                    cell.number_format = "yyyy-mm-dd"
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        for ci, h in enumerate(headers, 1):
            ml = max([len(str(h))] + [len(str(r.get(h, ""))) for r in rows])
            ws.column_dimensions[get_column_letter(ci)].width = min(ml + 2, 50)
        wb.save(path)


def render(template, row, extra=None):
    """Simple {var} substitution from row fields + extra dict."""
    out = template
    src = dict(row)
    if extra:
        src.update(extra)
    for k, v in src.items():
        if k is None:
            continue
        v = str(v or "").strip()
        out = out.replace("{" + str(k).lower() + "}", v)
        out = out.replace("{" + str(k) + "}", v)
    return out


def imap_append_to_folder(imap_host, user, pwd, folder, raw_message):
    """Save a sent copy into a custom IMAP folder (creates folder if missing)."""
    with imaplib.IMAP4_SSL(imap_host, 993, timeout=15) as imap:
        imap.login(user, pwd)
        # Try to create folder (no-op if exists)
        try:
            imap.create(folder)
        except Exception:
            pass
        # APPEND requires \Seen flag and a date — pass current time
        date_str = imaplib.Time2Internaldate(time.time())
        imap.append(folder, "\\Seen", date_str, raw_message.encode("utf-8"))


def _strip_html_to_plain(html_body):
    """Convert minimal HTML (just <a> tags) back to plain text with URL in parens.
    Used to generate the text/plain version of multipart messages."""
    plain = re.sub(r'<a\s+href="([^"]+)">([^<]+)</a>', r'\2 (\1)', html_body)
    # Strip any leftover tags (defensive)
    plain = re.sub(r'<[^>]+>', '', plain)
    return plain


def _wrap_html(body_with_anchors):
    """Wrap HTML body in minimal email-friendly HTML structure.
    Convert line breaks to <br> while preserving any existing <a> tags."""
    # Preserve plain-text feel: monospace-ish, no fancy styling.
    # Replace double-newlines with paragraph breaks, single newlines with <br>.
    paragraphs = body_with_anchors.split("\n\n")
    html_paragraphs = []
    for p in paragraphs:
        # Inside a paragraph, single newlines → <br>
        p_html = p.replace("\n", "<br>\n")
        html_paragraphs.append(f"<p>{p_html}</p>")
    html = "\n".join(html_paragraphs)
    return f"""<!DOCTYPE html><html><head><meta charset="utf-8"></head><body style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif; font-size: 14px; line-height: 1.5; color: #222;">
{html}
</body></html>"""


def build_message(from_name, from_addr, reply_to, to_addr, subject, body_with_anchors):
    """Build a multipart/alternative message with both text and HTML parts.
    `body_with_anchors` may contain <a href> tags; the plain-text version is
    auto-derived by converting anchors to 'label (url)' format."""
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = formataddr((from_name, from_addr)) if from_name else from_addr
    msg["To"] = to_addr
    msg["Date"] = formatdate(localtime=True)
    msg["Message-ID"] = make_msgid(domain=from_addr.split("@")[-1])
    if reply_to:
        msg["Reply-To"] = reply_to

    plain_body = _strip_html_to_plain(body_with_anchors)
    html_body = _wrap_html(body_with_anchors)

    # IMPORTANT: per RFC, plain MUST come first so non-HTML clients see it.
    msg.attach(MIMEText(plain_body, "plain", "utf-8"))
    msg.attach(MIMEText(html_body, "html", "utf-8"))
    return msg


def log_outcome(log_path, row, outcome, error=""):
    log_path.parent.mkdir(parents=True, exist_ok=True)
    new_file = not log_path.exists()
    with open(log_path, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if new_file:
            w.writerow(["timestamp", "email", "name", "tier", "outcome", "error"])
        w.writerow([datetime.now().isoformat(timespec="seconds"), row["Email"], row["Name"], row.get("Tier", ""), outcome, error])


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default="output/MASTER_youtube_outreach.xlsx")
    parser.add_argument("--config", default="outreach_config.yaml")
    parser.add_argument("--send", action="store_true", help="actually send (default is dry-run)")
    parser.add_argument("--limit", type=int, default=None, help="max emails this run")
    parser.add_argument("--tier", action="append", help="only this tier(s)")
    parser.add_argument("--min-fit", type=int, default=None, help="override min fit_score")
    parser.add_argument("--stage", choices=["welcome", "offer"], default="welcome",
                        help="welcome (cold intro, default) or offer (program info after reply)")
    args = parser.parse_args()

    base = Path(__file__).parent
    source = Path(args.source)
    if not source.is_absolute():
        source = base / source
    cfg = yaml.safe_load((base / args.config).read_text(encoding="utf-8"))

    sender_cfg = cfg.get("sender", {})
    filter_cfg = cfg.get("filters", {})

    if args.stage == "offer":
        offer_cfg = cfg.get("offer", {})
        subject_pool = offer_cfg.get("subjects", [])
        body_tpl = offer_cfg.get("body", "")
        intros = {}  # no per-tier intros for offer stage
        send_to_statuses = set(filter_cfg.get("send_offer_statuses", ["Replied"]))
        skip_statuses_set = None  # use inclusive filter instead
        new_status_after = "Sent program info"
    else:
        # Subject can be single string `subject:` or list `subjects:` (rotated)
        if "subjects" in cfg:
            subject_pool = cfg["subjects"]
            if isinstance(subject_pool, str):
                subject_pool = [subject_pool]
        else:
            subject_pool = [cfg.get("subject", "Hello {Name}")]
        body_tpl = cfg.get("body", "")
        intros = cfg.get("intros", {})
        send_to_statuses = None
        skip_statuses_set = set(filter_cfg.get("skip_statuses_welcome",
                                              filter_cfg.get("skip_statuses", [])))
        new_status_after = "Sent intro"

    min_fit = args.min_fit if args.min_fit is not None else filter_cfg.get("min_fit_score", 3)
    daily_limit = args.limit or sender_cfg.get("daily_limit", 50)
    min_delay = sender_cfg.get("min_delay_seconds", 30)
    max_delay = sender_cfg.get("max_delay_seconds", 60)
    imap_folder = sender_cfg.get("imap_folder", "")

    # SANITY CHECK: did MASTER lose its Status column data since last send?
    # Catches the bug that caused 25 duplicate sends on 26.05.
    if args.send and "MASTER_youtube_outreach" in str(source):
        try:
            from master_guard import sanity_check_status, backup_master
            ok, msg = sanity_check_status(tolerance_pct=30.0)
            print(f"[sanity] {msg}")
            if not ok:
                print(msg, file=sys.stderr)
                sys.exit(2)
            # Backup before any write
            bp = backup_master(reason=f"before-send-{args.stage}")
            if bp: print(f"[backup] saved {bp.name}")
        except ImportError:
            print("[warn] master_guard module missing — sending without safeguards", file=sys.stderr)

    rows, headers, fmt = load_rows(source)
    print(f"Loaded {len(rows)} rows from {source.name}")

    # Load already-in-CRM emails (from Google Sheet sync) — never re-send to these
    crm_emails_path = base / "output" / "crm_emails.csv"
    crm_emails = set()
    if crm_emails_path.exists():
        with open(crm_emails_path, encoding="utf-8") as f:
            for r in csv.DictReader(f):
                if r.get("email"):
                    crm_emails.add(r["email"].lower().strip())
        if crm_emails:
            print(f"  ({len(crm_emails)} contacts already in Google Sheet CRM — will skip)")

    # Filter candidates
    eligible = []
    excluded_crm = 0
    for r in rows:
        if not r.get("Email", "").strip():
            continue
        if r["Email"].lower().strip() in crm_emails:
            excluded_crm += 1
            continue
        status = (r.get("Status") or "").strip()
        if send_to_statuses is not None:
            # offer stage: include only matching statuses
            if status not in send_to_statuses:
                continue
        else:
            # welcome stage: exclude already-contacted / dead
            if status in (skip_statuses_set or set()):
                continue
        try:
            fit = int(r.get("fit_score") or 0)
        except (TypeError, ValueError):
            fit = 0
        if fit < min_fit:
            continue
        tier = r.get("Tier", "").strip().upper()
        if args.tier and tier not in {t.upper() for t in args.tier}:
            continue
        if args.stage == "welcome" and tier not in intros:
            print(f"  ! no intro for tier {tier!r} — skipping {r['Name']}", file=sys.stderr)
            continue
        eligible.append(r)

    print(f"Eligible (fit≥{min_fit}, not contacted): {len(eligible)}")
    if not eligible:
        print("Nothing to send.")
        return

    # Sort: Tier ascending (A→B→C→D), then fit_score desc, then Followers desc.
    # Picks highest-fit contacts first instead of relying on row order in xlsx.
    def _sort_key(r):
        try: fit = int(r.get("fit_score") or 0)
        except (TypeError, ValueError): fit = 0
        try: followers = int(r.get("Followers") or 0)
        except (TypeError, ValueError): followers = 0
        tier = (r.get("Tier", "") or "").strip().upper()
        return (tier, -fit, -followers)
    eligible.sort(key=_sort_key)

    to_send = eligible[:daily_limit]
    print(f"Will process: {len(to_send)} (daily_limit={daily_limit})")
    print(f"Mode: {'LIVE SEND ⚡' if args.send else 'DRY-RUN (no emails will be sent)'}")
    print("=" * 70)

    from_name = os.environ.get("FROM_NAME", "")
    from_addr = os.environ["SMTP_USER"]
    reply_to = os.environ.get("REPLY_TO", from_addr)
    host = os.environ["SMTP_HOST"]
    port = int(os.environ["SMTP_PORT"])
    pwd = os.environ["SMTP_PASSWORD"]

    log_path = base / "output" / "outreach_log.csv"

    def _connect_smtp(retries=3):
        """Connect to SMTP with retry on transient timeouts. Tries 465 SSL first,
        falls back to 587 STARTTLS if SSL handshake repeatedly fails."""
        last_err = None
        for attempt in range(1, retries + 1):
            try:
                ctx = ssl.create_default_context()
                s = smtplib.SMTP_SSL(host, port, context=ctx, timeout=20)
                s.login(from_addr, pwd)
                return s
            except (TimeoutError, OSError, smtplib.SMTPException) as e:
                last_err = e
                wait = 5 * attempt
                print(f"  ! SMTP_SSL attempt {attempt}/{retries} failed: {type(e).__name__}: {e}. "
                      f"Retrying in {wait}s...", file=sys.stderr)
                time.sleep(wait)
        # SSL exhausted → try STARTTLS on 587
        print("  ! Falling back to STARTTLS on port 587...", file=sys.stderr)
        try:
            s = smtplib.SMTP(host, 587, timeout=20)
            s.ehlo()
            s.starttls(context=ssl.create_default_context())
            s.ehlo()
            s.login(from_addr, pwd)
            return s
        except Exception as e:
            raise RuntimeError(f"SMTP connect failed on both 465/SSL and 587/STARTTLS. Last: {last_err}; STARTTLS: {e}")

    # SMTP connection (only if --send)
    smtp = None
    if args.send:
        smtp = _connect_smtp()
        print(f"✓ SMTP connected as {from_addr}")

    sent_count, failed_count = 0, 0
    imap_host = host.replace("smtp.", "imap.") if "smtp." in host else "imap.yandex.ru"
    for i, r in enumerate(to_send, 1):
        tier = r["Tier"].upper()
        # Smart-extract first name from channel name / email.
        # If extraction fails — leave Name empty and post-process greeting to
        # just "Hey!" (no brand-name addressing).
        extracted = smart_extract_name(r.get("Name", ""), r.get("Email", ""))
        r_for_render = dict(r)
        r_for_render["Name"] = extracted  # empty string if extraction failed
        # Rotate subject round-robin across the pool
        subject_tpl = subject_pool[(i - 1) % len(subject_pool)]
        # Render: intro first (it has {Platform} etc.), then plug into body
        if args.stage == "welcome":
            intro = render(intros[tier], r_for_render)
            body = render(body_tpl, r_for_render, extra={"intro": intro})
        else:
            body = render(body_tpl, r_for_render)
        subject = render(subject_tpl, r_for_render)

        # Clean up artifacts when {Name} was empty (smart-extract failed)
        # "Hey !" → "Hey!", "Hey  —" → "Hey —", "Hey  T" → "Hey T", etc.
        if not extracted:
            body = re.sub(r"Hey\s+!", "Hey!", body)
            body = re.sub(r"Hey\s+,", "Hey,", body)
            body = re.sub(r"Hey\s{2,}", "Hey ", body)
            subject = re.sub(r"Hey\s+—", "Hey —", subject)
            subject = re.sub(r"Hey\s+!", "Hey!", subject)
            subject = re.sub(r"Hey\s{2,}", "Hey ", subject)
            # Subject "Hey  — Timestripe..." → strip leading "Hey — " if subject starts there
            subject = re.sub(r"^Hey\s*—\s*", "", subject)

        print(f"\n[{i}/{len(to_send)}] {r['Name']} <{r['Email']}>  Tier {tier}  fit={r.get('fit_score','?')}")
        print(f"  Subject: {subject}")
        if not args.send:
            print(f"  --- body preview (first 10 lines) ---")
            for line in body.split("\n")[:10]:
                print(f"    {line}")
            print(f"  ✓ DRY-RUN (would send)")
            continue

        msg = build_message(from_name, from_addr, reply_to, r["Email"], subject, body)
        raw = msg.as_string()
        # Try sending with up to 2 retries on transient errors (reconnect SMTP if needed)
        send_ok, send_err = False, None
        for attempt in range(1, 3):
            try:
                smtp.sendmail(from_addr, [r["Email"]], raw)
                send_ok = True
                break
            except (smtplib.SMTPServerDisconnected, smtplib.SMTPConnectError,
                    TimeoutError, OSError) as e:
                send_err = e
                print(f"  ! send attempt {attempt} failed ({type(e).__name__}); reconnecting...")
                try: smtp.quit()
                except Exception: pass
                try:
                    smtp = _connect_smtp(retries=2)
                except Exception as recon_err:
                    send_err = recon_err
                    break
            except smtplib.SMTPRecipientsRefused as e:
                send_err = e
                break  # not retryable (bad recipient)

        if send_ok:
            if imap_folder:
                try:
                    imap_append_to_folder(imap_host, from_addr, pwd, imap_folder, raw)
                except Exception as e:
                    print(f"  ! IMAP copy to {imap_folder!r} failed: {e}")
            today = datetime.now().strftime("%Y-%m-%d")
            r["Status"] = new_status_after
            if args.stage == "welcome":
                r["First Contact Date"] = today
            r["Last Follow-up"] = today
            log_outcome(log_path, r, "sent")
            print(f"  ✓ sent" + (f" + saved to {imap_folder}" if imap_folder else ""))
            sent_count += 1
        else:
            log_outcome(log_path, r, "failed", str(send_err))
            print(f"  ✗ FAILED: {send_err}")
            failed_count += 1

        # Save progress after each send so we don't lose state on Ctrl+C
        save_rows(source, rows, headers, fmt)

        # Rate limit
        if i < len(to_send):
            delay = random.uniform(min_delay, max_delay)
            print(f"  ... sleeping {delay:.0f}s")
            time.sleep(delay)

    if smtp:
        smtp.quit()
    print("\n" + "=" * 70)
    print(f"Done. Sent: {sent_count}  Failed: {failed_count}  Dry-run: {len(to_send)-sent_count-failed_count}")
    if args.send:
        print(f"Source updated:  {source}")
        print(f"Send log:        {log_path}")
        # Push status changes to Google Sheet + repaint
        if sent_count > 0:
            import subprocess as _sp
            try:
                print("\nSyncing status to Google Sheet...")
                _sp.run([sys.executable, str(base / "sheet_status_sync.py"), "--paint"],
                        cwd=str(base), check=False, timeout=120)
            except Exception as e:
                print(f"! Sheet sync failed: {e} (master is still up to date)")


if __name__ == "__main__":
    main()
