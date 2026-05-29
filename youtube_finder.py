#!/usr/bin/env python3
"""
YouTube influencer finder for TimeStripe outreach.

Searches YouTube tier-by-tier (A/B/C/D), each tier has its own keywords +
subscriber range. Outputs a CSV with columns matching the Google Sheet CRM
(tab "Vlad"). Channels WITHOUT a public email are excluded from the main CSV
and dumped to a separate "skipped" file for manual lookup.

Uses YouTube Data API v3 (free, 10,000 quota units/day):
  - search.list = 100 units per call
  - channels.list = 1 unit per call
"""
import argparse
import csv
import os
import re
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

import yaml
from dotenv import load_dotenv
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

API_KEY = os.environ.get("YOUTUBE_API_KEY")
if not API_KEY:
    print("ERROR: set YOUTUBE_API_KEY in .env (see .env.example)", file=sys.stderr)
    sys.exit(1)

# Email regex — slightly stricter to avoid junk matches like "blah@1x.png"
EMAIL_RE = re.compile(r"\b[\w.+-]+@[A-Za-z0-9-]+\.[A-Za-z]{2,}(?:\.[A-Za-z]{2,})?\b")
CYRILLIC_RE = re.compile(r"[А-яЁё]")
SOCIAL_RES = {
    "instagram": re.compile(r"(?:instagram\.com|instagr\.am)/[\w._-]+", re.IGNORECASE),
    "tiktok": re.compile(r"tiktok\.com/@[\w._-]+", re.IGNORECASE),
    "telegram": re.compile(r"(?:t\.me|telegram\.me)/[\w_]+", re.IGNORECASE),
    "twitter": re.compile(r"(?:twitter\.com|x\.com)/[\w_]+", re.IGNORECASE),
    "linktree": re.compile(r"(?:linktr\.ee|linktree\.com|beacons\.ai|stan\.store|bio\.link)/[\w._-]+", re.IGNORECASE),
}

# Fit-scoring — channels that look like real Timestripe users (productivity vibe)
PRODUCTIVITY_TERMS_RE = re.compile(
    r"\b(plann?(ing|er)|productiv|journal|vision board|goal[s]?|morning routine|"
    r"intentional|habit[s]?|discipline|life reset|sunday reset|study with me|"
    r"study vlog|day in my life|aesthetic|self growth|self improvement|"
    r"organi[sz]ation|deep work|focus|mindful|reflection|review|bullet journal|"
    r"todo|to-do|to do list|calendar|notion|workflow|second brain|gtd|kanban)\b",
    re.IGNORECASE,
)
MEMORIAL_RE = re.compile(r"RIP|tribute|memorial|in heaven|👼|🕊", re.IGNORECASE)
ESOTERIC_RE = re.compile(
    r"\b(astrology|tarot|prediction|prophecy|horoscope|breaking news|news update|crypto signals?)\b",
    re.IGNORECASE,
)


def compute_fit_score(name, description, subs, video_count, last_upload_days_ago=None):
    """Score a channel by how Timestripe-target it looks. +5 great, 0 neutral, -3 junk."""
    score = 0
    flags = []
    text = (name or "") + " " + (description or "")

    # +: productivity terms
    matches = PRODUCTIVITY_TERMS_RE.findall(text)
    if matches:
        bonus = min(len(matches) * 2, 8)
        score += bonus
        flags.append(f"prod={len(matches)}")

    # -: spammy videos:subs ratio
    ratio = video_count / max(subs, 1)
    if ratio > 0.5:
        score -= 6; flags.append(f"spam_ratio={ratio:.2f}")
    elif ratio > 0.3:
        score -= 3; flags.append(f"high_ratio={ratio:.2f}")

    # +/-: recency
    if last_upload_days_ago is not None:
        if last_upload_days_ago < 30: score += 3
        elif last_upload_days_ago < 90: score += 1
        elif last_upload_days_ago > 365: score -= 4; flags.append(f"dead_{last_upload_days_ago}d")
        flags.append(f"last_{last_upload_days_ago}d")

    # -: red flags
    if MEMORIAL_RE.search(text):
        score -= 10; flags.append("MEMORIAL")
    if ESOTERIC_RE.search(text):
        score -= 4; flags.append("ESOTERIC")

    return score, "; ".join(flags)


# Columns matching the user's Google Sheet (tab "Vlad").
CRM_COLUMNS = [
    "Name", "Email", "Platform", "URL", "Followers", "Tier", "Niche",
    "Status", "First Contact Date", "Last Follow-up", "Notes",
    "Program Type", "Videos Published", "Total Views",
    "Payout Owed", "Payout Sent", "ROI",
    "fit_score", "fit_flags",
]


def chunked(seq, n):
    seq = list(seq)
    for i in range(0, len(seq), n):
        yield seq[i:i + n]


def search_channels_by_keyword(yt, keyword, max_results, published_after, region_code, relevance_language):
    channel_ids = set()
    page_token = None
    fetched = 0
    while fetched < max_results:
        params = {
            "q": keyword,
            "part": "snippet",
            "type": "video",
            "order": "date",
            "maxResults": min(50, max_results - fetched),
            "pageToken": page_token,
        }
        if published_after:
            params["publishedAfter"] = published_after
        if region_code:
            params["regionCode"] = region_code
        if relevance_language:
            params["relevanceLanguage"] = relevance_language
        try:
            resp = yt.search().list(**params).execute()
        except HttpError as e:
            print(f"  ! search error for {keyword!r}: {e}", file=sys.stderr)
            break
        items = resp.get("items", [])
        for item in items:
            channel_ids.add(item["snippet"]["channelId"])
        fetched += len(items)
        page_token = resp.get("nextPageToken")
        if not page_token or not items:
            break
    return channel_ids


def get_channel_details(yt, channel_ids, fetch_last_upload=True):
    results = {}
    for batch in chunked(channel_ids, 50):
        try:
            resp = yt.channels().list(
                part="snippet,statistics,brandingSettings,contentDetails",
                id=",".join(batch),
                maxResults=50,
            ).execute()
        except HttpError as e:
            print(f"  ! channels.list error: {e}", file=sys.stderr)
            continue
        for item in resp.get("items", []):
            stats = item.get("statistics", {})
            snippet = item.get("snippet", {})
            branding = item.get("brandingSettings", {}).get("channel", {})
            hidden = bool(stats.get("hiddenSubscriberCount"))
            custom = snippet.get("customUrl", "")
            if custom and custom.startswith("@"):
                url = f"https://www.youtube.com/{custom}"
            elif custom:
                url = f"https://www.youtube.com/@{custom.lstrip('@')}"
            else:
                url = f"https://www.youtube.com/channel/{item['id']}"
            results[item["id"]] = {
                "channel_id": item["id"],
                "name": snippet.get("title", ""),
                "subscribers": None if hidden else int(stats.get("subscriberCount", 0) or 0),
                "subs_hidden": hidden,
                "video_count": int(stats.get("videoCount", 0) or 0),
                "view_count": int(stats.get("viewCount", 0) or 0),
                "country": snippet.get("country", "") or branding.get("country", ""),
                "description": (snippet.get("description") or "")[:2000],
                "url": url,
                "uploads_playlist": item.get("contentDetails", {}).get("relatedPlaylists", {}).get("uploads"),
                "last_upload_days_ago": None,
            }

    # Fetch last-upload date per channel (1 quota unit each — cheap)
    if fetch_last_upload:
        for cid, d in results.items():
            pid = d.get("uploads_playlist")
            if not pid:
                continue
            try:
                r = yt.playlistItems().list(part="snippet", playlistId=pid, maxResults=1).execute()
                items = r.get("items", [])
                if items:
                    pub = items[0]["snippet"]["publishedAt"]
                    last_dt = datetime.fromisoformat(pub.replace("Z", "+00:00"))
                    d["last_upload_days_ago"] = (datetime.now(timezone.utc) - last_dt).days
            except HttpError:
                pass

    return results


def extract_contacts(text):
    text = text or ""
    emails = sorted(set(m.strip(".") for m in EMAIL_RE.findall(text)))
    # Drop obviously bogus matches
    emails = [e for e in emails if not e.lower().endswith((".png", ".jpg", ".jpeg", ".gif", ".svg"))]
    socials = {}
    for name, pat in SOCIAL_RES.items():
        matches = sorted(set(pat.findall(text)))
        socials[name] = matches
    return emails, socials


def build_notes(emails, socials, matched_kws):
    """Compose the Notes cell — secondary emails, extra socials, top keywords."""
    parts = []
    if len(emails) > 1:
        parts.append("alt emails: " + ", ".join(emails[1:]))
    for label, key in [("IG", "instagram"), ("TT", "tiktok"), ("TG", "telegram"), ("X", "twitter"), ("LINK", "linktree")]:
        urls = socials.get(key) or []
        if urls:
            # Use first match; prepend https:// if it's a bare domain match
            u = urls[0]
            if not u.startswith("http"):
                u = "https://" + u
            parts.append(f"{label}: {u}")
    if matched_kws:
        top = sorted(matched_kws)[:3]
        parts.append("kw: " + ", ".join(top))
    return " | ".join(parts)


def run_tier(yt, tier, defaults):
    name = tier["name"]
    niche_label = tier.get("niche_label", name)
    min_subs = tier.get("min_subscribers", defaults.get("min_subscribers", 0))
    max_subs = tier.get("max_subscribers", defaults.get("max_subscribers", 1_000_000))
    min_videos = tier.get("min_videos", defaults.get("min_videos", 1))
    results_per_kw = tier.get("results_per_keyword", defaults.get("results_per_keyword", 50))
    region = tier.get("region_code", defaults.get("region_code"))
    lang = tier.get("relevance_language", defaults.get("relevance_language"))
    days = tier.get("days_back", defaults.get("days_back"))
    skip_cyrillic = tier.get("skip_cyrillic_names", defaults.get("skip_cyrillic_names", False))
    min_fit = tier.get("min_fit_score", defaults.get("min_fit_score", -999))
    published_after = None
    if days:
        published_after = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")

    print(f"\n=== Tier {name} ({niche_label}) | subs {min_subs}-{max_subs} | {len(tier['keywords'])} keywords ===")

    keyword_map = {}
    for i, kw in enumerate(tier["keywords"], 1):
        print(f"  [{i}/{len(tier['keywords'])}] {kw!r}", end=" ... ", flush=True)
        ids = search_channels_by_keyword(yt, kw, results_per_kw, published_after, region, lang)
        print(f"{len(ids)} channels")
        for cid in ids:
            keyword_map.setdefault(cid, set()).add(kw)
        time.sleep(0.15)

    all_ids = list(keyword_map.keys())
    print(f"  Unique channels in tier: {len(all_ids)}. Fetching stats...")
    details = get_channel_details(yt, all_ids)

    matched, skipped = [], []
    dropped_cyrillic = 0
    dropped_low_fit = 0
    for cid, d in details.items():
        if d["subs_hidden"] or d["subscribers"] is None:
            continue
        if not (min_subs <= d["subscribers"] <= max_subs):
            continue
        if d["video_count"] < min_videos:
            continue
        if skip_cyrillic and CYRILLIC_RE.search(d["name"]):
            dropped_cyrillic += 1
            continue

        fit_score, fit_flags = compute_fit_score(
            d["name"], d["description"], d["subscribers"], d["video_count"],
            d.get("last_upload_days_ago"),
        )
        if fit_score < min_fit:
            dropped_low_fit += 1
            continue

        emails, socials = extract_contacts(d["description"])
        matched_kws = keyword_map.get(cid, set())
        notes = build_notes(emails, socials, matched_kws)

        base = {
            "Name": d["name"],
            "Email": emails[0] if emails else "",
            "Platform": "YouTube",
            "URL": d["url"],
            "Followers": d["subscribers"],
            "Tier": name.split("_")[0].upper(),
            "Niche": niche_label,
            "Status": "",
            "First Contact Date": "",
            "Last Follow-up": "",
            "Notes": notes,
            "Program Type": "",
            "Videos Published": d["video_count"],
            "Total Views": d["view_count"],
            "Payout Owed": "",
            "Payout Sent": "",
            "ROI": "",
            "fit_score": fit_score,
            "fit_flags": fit_flags,
            # Internal-only:
            "_channel_id": cid,
            "_country": d["country"],
        }
        if emails:
            matched.append(base)
        else:
            skipped.append(base)

    msg = f"  → {len(matched)} with email (kept), {len(skipped)} no-email (skipped)"
    if skip_cyrillic: msg += f", {dropped_cyrillic} cyrillic"
    if min_fit > -999: msg += f", {dropped_low_fit} low-fit (<{min_fit})"
    print(msg)
    return matched, skipped


def write_csv(path, rows, include_internal=False):
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(CRM_COLUMNS)
    if include_internal:
        fields += ["_channel_id", "_country"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def write_xlsx(path, rows, include_internal=False, sheet_title="Sheet1"):
    """Same data as CSV but as proper .xlsx — Excel/Google Sheets-friendly."""
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(CRM_COLUMNS)
    if include_internal:
        fields += ["_channel_id", "_country"]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel limit

    ws.append(fields)
    for r in rows:
        ws.append([r.get(f, "") for f in fields])

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze top row + autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Auto-width per column (capped at 50)
    for col_idx, field in enumerate(fields, 1):
        max_len = len(field)
        for r in rows:
            v = r.get(field, "")
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    # Format numeric columns with thousands separator
    numeric_cols = {"Followers", "Videos Published", "Total Views"}
    for col_idx, field in enumerate(fields, 1):
        if field in numeric_cols:
            for cell in ws[get_column_letter(col_idx)][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Find YouTube influencers by tier for TimeStripe outreach.")
    parser.add_argument("--config", default="config.yaml")
    parser.add_argument("--out", default=None, help="output CSV path (main, email-only)")
    parser.add_argument("--skipped-out", default=None, help="output CSV path (no-email rows, for manual lookup)")
    parser.add_argument("--tier", action="append", help="run only specific tier name(s); repeatable")
    parser.add_argument("--region", default=None, help="override ISO country code")
    parser.add_argument("--lang", default=None, help="override relevance language")
    parser.add_argument("--days", type=int, default=None, help="override days_back")
    parser.add_argument("--seen-file", default="output/seen_channels.csv",
                        help="path to journal of channel IDs already discovered in previous runs")
    parser.add_argument("--full", action="store_true",
                        help="ignore seen-file, re-fetch everything (default is delta mode — only new)")
    args = parser.parse_args()

    cfg_path = Path(args.config)
    if not cfg_path.is_absolute():
        cfg_path = Path(__file__).parent / cfg_path
    cfg = yaml.safe_load(cfg_path.read_text(encoding="utf-8"))

    defaults = cfg.get("defaults", {}) or {}
    if args.region:
        defaults["region_code"] = args.region
    if args.lang:
        defaults["relevance_language"] = args.lang
    if args.days is not None:
        defaults["days_back"] = args.days

    tiers = cfg["tiers"]
    if args.tier:
        wanted = set(args.tier)
        tiers = [t for t in tiers if t["name"] in wanted]
        if not tiers:
            print(f"No tiers matched filter {wanted}. Available: {[t['name'] for t in cfg['tiers']]}", file=sys.stderr)
            sys.exit(1)

    yt = build("youtube", "v3", developerKey=API_KEY)

    # Load journal of channels discovered in previous runs (cross-run dedup)
    seen_file = Path(args.seen_file)
    if not seen_file.is_absolute():
        seen_file = Path(__file__).parent / seen_file
    previously_seen = set()
    if seen_file.exists() and not args.full:
        with open(seen_file, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                previously_seen.add(row["channel_id"])
        print(f"Loaded {len(previously_seen)} previously-seen channels from {seen_file.name}")
        print("Mode: DELTA — only NEW channels will be written to output")
    else:
        if args.full:
            print("Mode: FULL — re-fetching everything (will overwrite seen-file)")
        else:
            print(f"Mode: DELTA (no prior seen-file at {seen_file.name} — first run)")

    matched_rows, skipped_rows = [], []
    seen_emails = set()       # dedup primary CSV by email (within this run)
    seen_channel_ids = set()  # dedup skipped CSV by channel (within this run)
    newly_seen = set()        # channels discovered in this run (for journal update)
    cross_run_dedup_count = 0

    for tier in tiers:
        m, s = run_tier(yt, tier, defaults)
        for r in m:
            cid = r["_channel_id"]
            newly_seen.add(cid)
            # Cross-run dedup: skip if seen in prior runs
            if cid in previously_seen:
                cross_run_dedup_count += 1
                continue
            email_key = r["Email"].lower().strip()
            if email_key in seen_emails:
                continue
            seen_emails.add(email_key)
            matched_rows.append(r)
        for r in s:
            cid = r["_channel_id"]
            newly_seen.add(cid)
            if cid in previously_seen:
                continue
            if cid in seen_channel_ids:
                continue
            seen_channel_ids.add(cid)
            skipped_rows.append(r)

    if cross_run_dedup_count and not args.full:
        print(f"\n  Skipped {cross_run_dedup_count} channels already seen in prior runs (delta mode)")

    # Sort by tier, then fit_score desc, then followers desc — best leads first
    matched_rows.sort(key=lambda r: (r["Tier"], -r.get("fit_score", 0), -r["Followers"]))
    skipped_rows.sort(key=lambda r: (r["Tier"], -r.get("fit_score", 0), -r["Followers"]))

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out_main = Path(args.out) if args.out else Path(__file__).parent / f"output/youtube_{ts}.csv"
    out_skip = Path(args.skipped_out) if args.skipped_out else Path(__file__).parent / f"output/youtube_skipped_no_email_{ts}.csv"

    write_csv(out_main, matched_rows)
    write_csv(out_skip, skipped_rows, include_internal=True)
    write_xlsx(out_main.with_suffix(".xlsx"), matched_rows, sheet_title="YouTube influencers")
    write_xlsx(out_skip.with_suffix(".xlsx"), skipped_rows, include_internal=True, sheet_title="Skipped (no email)")

    # Summary
    by_tier = {}
    for r in matched_rows:
        by_tier[r["Tier"]] = by_tier.get(r["Tier"], 0) + 1
    print("\n=== Summary (with email — going into CRM) ===")
    for t in sorted(by_tier):
        print(f"  Tier {t}: {by_tier[t]}")
    print(f"Total kept: {len(matched_rows)}  |  Skipped (no email): {len(skipped_rows)}")
    print(f"→ Main CSV:  {out_main}")
    print(f"→ Main XLSX: {out_main.with_suffix('.xlsx')}")
    print(f"→ Skipped CSV:  {out_skip}")
    print(f"→ Skipped XLSX: {out_skip.with_suffix('.xlsx')}")

    # Update seen-file journal (UNION of previously_seen + newly_seen)
    seen_file.parent.mkdir(parents=True, exist_ok=True)
    union = previously_seen | newly_seen
    with open(seen_file, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["channel_id", "first_seen"])
        w.writeheader()
        today = datetime.now().strftime("%Y-%m-%d")
        for cid in sorted(union):
            w.writerow({"channel_id": cid, "first_seen": today if cid not in previously_seen else ""})
    print(f"→ Seen journal: {seen_file} ({len(union)} channel IDs total)")


if __name__ == "__main__":
    main()
