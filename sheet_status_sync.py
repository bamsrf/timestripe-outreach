#!/usr/bin/env python3
"""
Sync status fields MASTER xlsx → Google Sheet.

For each row in MASTER, find matching row in Sheet by Email and copy over:
- Status
- First Contact Date
- Last Follow-up

Idempotent — only writes cells that actually differ.

Usage:
    python sheet_status_sync.py                 # sync all MASTER → Sheet
    python sheet_status_sync.py --paint         # also repaint colors after
    python sheet_status_sync.py --emails a@b.com,c@d.com  # only these
"""
import argparse
import os
import subprocess
import sys
from datetime import date as _date, datetime
from pathlib import Path

import gspread
from dotenv import load_dotenv
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook

ROOT = Path(__file__).parent
load_dotenv(ROOT / ".env")

DEFAULT_CREDS_PATH = Path.home() / ".config" / "timestripe-gsheets.json"


def date_to_sheet_format(value):
    """Coerce date-ish value to DD.MM.YY format used in the user's sheet."""
    if not value or not str(value).strip():
        return ""
    if isinstance(value, _date):
        return value.strftime("%d.%m.%y")
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date().strftime("%d.%m.%y")
        except ValueError:
            continue
    return s  # leave as-is if unparseable


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--master", default="output/MASTER_youtube_outreach.xlsx")
    parser.add_argument("--tab", default="Vlad")
    parser.add_argument("--emails", default=None,
                        help="comma-separated emails to sync (default: all in MASTER)")
    parser.add_argument("--paint", action="store_true", help="run paint_sheet.py after sync")
    args = parser.parse_args()

    # Load MASTER
    master_path = ROOT / args.master if not Path(args.master).is_absolute() else Path(args.master)
    if not master_path.exists():
        sys.exit(f"ERROR: master not found at {master_path}")
    wb = load_workbook(master_path)
    ws_m = wb.active
    m_headers = [c.value for c in ws_m[1]]

    needed = ["Email", "Status", "First Contact Date", "Last Follow-up"]
    for h in needed:
        if h not in m_headers:
            sys.exit(f"ERROR: master missing column {h!r}")
    em_i = m_headers.index("Email")
    st_i = m_headers.index("Status")
    fcd_i = m_headers.index("First Contact Date")
    lfu_i = m_headers.index("Last Follow-up")

    # Build email → state map from MASTER (only rows with non-empty Status)
    master_state = {}
    for row in ws_m.iter_rows(min_row=2, values_only=True):
        em = (row[em_i] or "").strip().lower()
        if not em:
            continue
        status = (row[st_i] or "").strip()
        if not status:
            continue  # skip uncontacted — no info to sync
        master_state[em] = {
            "Status": status,
            "First Contact Date": date_to_sheet_format(row[fcd_i]),
            "Last Follow-up": date_to_sheet_format(row[lfu_i]),
        }

    # Filter by --emails if specified
    if args.emails:
        wanted = {e.strip().lower() for e in args.emails.split(",") if e.strip()}
        master_state = {e: s for e, s in master_state.items() if e in wanted}

    if not master_state:
        print("No master rows with Status to sync")
        return

    # Connect to sheet
    sheet_id = os.environ.get("GSHEETS_ID")
    if not sheet_id:
        sys.exit("ERROR: GSHEETS_ID missing in .env")
    creds_path = Path(os.environ.get("GSHEETS_CREDS", DEFAULT_CREDS_PATH))
    creds = Credentials.from_service_account_file(
        str(creds_path),
        scopes=["https://www.googleapis.com/auth/spreadsheets"],
    )
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)
    ws = sh.worksheet(args.tab)

    all_rows = ws.get_all_values()
    headers, header_row = None, None
    for i, row in enumerate(all_rows, start=1):
        if row and row[0].strip().lower() == "name":
            headers, header_row = row, i
            break
    if not header_row:
        sys.exit("ERROR: header row (Name in col A) not found in sheet")

    email_col = headers.index("Email") + 1
    status_col = headers.index("Status") + 1
    fcd_col = headers.index("First Contact Date") + 1
    lfu_col = headers.index("Last Follow-up") + 1

    updates = []
    changed_rows = []
    for row_idx, row in enumerate(all_rows, start=1):
        if row_idx <= header_row or len(row) <= status_col - 1:
            continue
        em = row[email_col - 1].lower().strip()
        if em not in master_state:
            continue
        state = master_state[em]
        current_status = row[status_col - 1] if len(row) >= status_col else ""
        current_fcd = row[fcd_col - 1] if len(row) >= fcd_col else ""
        current_lfu = row[lfu_col - 1] if len(row) >= lfu_col else ""

        row_changed = False
        if current_status != state["Status"]:
            updates.append({"range": gspread.utils.rowcol_to_a1(row_idx, status_col),
                            "values": [[state["Status"]]]})
            row_changed = True
        if state["First Contact Date"] and current_fcd != state["First Contact Date"]:
            updates.append({"range": gspread.utils.rowcol_to_a1(row_idx, fcd_col),
                            "values": [[state["First Contact Date"]]]})
            row_changed = True
        if state["Last Follow-up"] and current_lfu != state["Last Follow-up"]:
            updates.append({"range": gspread.utils.rowcol_to_a1(row_idx, lfu_col),
                            "values": [[state["Last Follow-up"]]]})
            row_changed = True
        if row_changed:
            changed_rows.append((row_idx, row[0], state["Status"]))

    if updates:
        ws.batch_update(updates, value_input_option="USER_ENTERED")
        print(f"✓ Synced {len(changed_rows)} rows ({len(updates)} cells):")
        for idx, name, status in changed_rows[:20]:
            print(f"   Row {idx}: {name[:35]:<35} → {status}")
        if len(changed_rows) > 20:
            print(f"   ... +{len(changed_rows)-20} more")
    else:
        print("Sheet already in sync with MASTER")

    if args.paint:
        print("\nRepainting...")
        subprocess.run([sys.executable, str(ROOT / "paint_sheet.py")], cwd=str(ROOT))


if __name__ == "__main__":
    main()
