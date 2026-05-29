#!/usr/bin/env python3
"""
TimeStripe Outreach — admin dashboard.

Run:
    cd /Users/vladislavrumancev/Desktop/Cursor/Timestripe
    source .venv/bin/activate
    streamlit run admin.py

Opens in browser at http://localhost:8501.

Pages:
  📊 Overview      KPI tiles + funnel chart + breakdowns
  👥 Contacts      Searchable table, inline Status editing
  📬 Send Queue    Preview next N to send, fire `outreach_sender.py --send`
  📝 Templates     View welcome + offer templates
  📜 Send Log      Full history from outreach_log.csv
"""
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import yaml

ROOT = Path(__file__).parent
MASTER_FILE = ROOT / "output" / "MASTER_youtube_outreach.xlsx"
LOG_FILE = ROOT / "output" / "outreach_log.csv"
CONFIG_FILE = ROOT / "outreach_config.yaml"

# Status progression — used for funnel and color coding
FUNNEL_STAGES = [
    ("Eligible (not contacted)", lambda s: not s or s == "Not contacted"),
    ("Welcome sent", lambda s: s == "Sent intro"),
    ("Replied", lambda s: s == "Replied"),
    ("Offer sent", lambda s: s == "Sent program info"),
    ("Negotiating", lambda s: s == "Wants money for posts"),
    ("Won (live collab)", lambda s: s == "Live collab"),
]
TERMINAL_NEGATIVE = {"Dead", "Do not contact"}

# ----------------------------------------------------------------------
# Data loading (cached for performance, refresh button clears cache)
# ----------------------------------------------------------------------

@st.cache_data(ttl=10)
def load_master() -> pd.DataFrame:
    if not MASTER_FILE.exists():
        return pd.DataFrame()
    df = pd.read_excel(MASTER_FILE)
    df["Status"] = df["Status"].fillna("").astype(str)
    for col in ("Followers", "Videos Published", "Total Views", "fit_score"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    return df


@st.cache_data(ttl=10)
def load_log() -> pd.DataFrame:
    if not LOG_FILE.exists():
        return pd.DataFrame(columns=["timestamp", "email", "name", "tier", "outcome", "error"])
    return pd.read_csv(LOG_FILE)


@st.cache_data(ttl=10)
def load_config() -> dict:
    if not CONFIG_FILE.exists():
        return {}
    return yaml.safe_load(CONFIG_FILE.read_text(encoding="utf-8"))


DATE_COLUMNS = {"First Contact Date", "Last Follow-up"}


def _coerce_date(value):
    from datetime import date as _date, datetime as _dt
    if isinstance(value, _date):
        return value
    if pd.isna(value) or not str(value).strip():
        return ""
    try:
        return _dt.strptime(str(value).strip(), "%Y-%m-%d").date()
    except ValueError:
        return value


def save_master(df: pd.DataFrame):
    """Save edited DF back to xlsx with proper date types in date columns."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # Backup current MASTER before overwriting (so we can recover from any bug)
    try:
        from master_guard import backup_master
        backup_master(reason="admin-save")
    except ImportError:
        pass

    wb = Workbook(); ws = wb.active; ws.title = "Master YT outreach"
    headers = list(df.columns)
    ws.append(headers)
    for _, row in df.iterrows():
        row_out = []
        for h in headers:
            v = row[h] if pd.notna(row[h]) else ""
            if h in DATE_COLUMNS:
                v = _coerce_date(v)
            row_out.append(v)
        ws.append(row_out)
    hf = Font(bold=True, color="FFFFFF"); hb = PatternFill("solid", fgColor="4F81BD")
    for cell in ws[1]:
        cell.font = hf; cell.fill = hb; cell.alignment = Alignment(horizontal="center")
    for ci, h in enumerate(headers, 1):
        if h in DATE_COLUMNS:
            for cell in ws[get_column_letter(ci)][1:]:
                cell.number_format = "yyyy-mm-dd"
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    for ci, h in enumerate(headers, 1):
        ml = max([len(str(h))] + [len(str(v)) for v in df[h].astype(str).head(200)])
        ws.column_dimensions[get_column_letter(ci)].width = min(ml + 2, 50)
    wb.save(MASTER_FILE)
    st.cache_data.clear()


def stage_of(status: str) -> str:
    s = (status or "").strip()
    if s in TERMINAL_NEGATIVE:
        return s
    for label, predicate in FUNNEL_STAGES:
        if predicate(s):
            return label
    return f"Other ({s})"


# ----------------------------------------------------------------------
# Pages
# ----------------------------------------------------------------------

def page_overview():
    st.title("📊 Outreach Overview")
    df = load_master()
    log = load_log()
    if df.empty:
        st.warning(f"No data — generate {MASTER_FILE.name} first.")
        return

    # KPI tiles
    total = len(df)
    sent = (df["Status"] == "Sent intro").sum() + (df["Status"] == "Sent program info").sum()
    pending = df.apply(lambda r: not r["Status"] or r["Status"] == "Not contacted", axis=1).sum()
    replied = (df["Status"] == "Replied").sum()
    won = (df["Status"] == "Live collab").sum()
    sends_today = 0
    if not log.empty and "timestamp" in log.columns:
        today = datetime.now().strftime("%Y-%m-%d")
        sends_today = log["timestamp"].astype(str).str.startswith(today).sum()

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Total contacts", total)
    c2.metric("Pending", pending)
    c3.metric("Sent (any)", sent)
    c4.metric("Replied", replied)
    c5.metric("Won", won)
    c6.metric("Sent today", sends_today)

    st.divider()

    # Funnel chart
    st.subheader("Funnel")
    stages, counts = [], []
    for label, predicate in FUNNEL_STAGES:
        stages.append(label)
        counts.append(int(df["Status"].apply(predicate).sum()))
    fig = go.Figure(go.Funnel(
        y=stages, x=counts,
        textposition="inside", textinfo="value+percent initial",
        marker={"color": ["#5B9BD5", "#70AD47", "#FFC000", "#ED7D31", "#A5A5A5", "#4472C4"]},
    ))
    fig.update_layout(height=400, margin=dict(l=20, r=20, t=20, b=20))
    st.plotly_chart(fig, use_container_width=True)

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("By Tier")
        tier_counts = df["Tier"].value_counts().reset_index()
        tier_counts.columns = ["Tier", "Count"]
        st.plotly_chart(
            px.bar(tier_counts, x="Tier", y="Count", color="Tier",
                   text="Count", height=300).update_layout(showlegend=False, margin=dict(t=10)),
            use_container_width=True,
        )
    with col2:
        st.subheader("By Status")
        status_counts = df["Status"].replace("", "Not contacted").value_counts().reset_index()
        status_counts.columns = ["Status", "Count"]
        st.plotly_chart(
            px.bar(status_counts, x="Count", y="Status", orientation="h",
                   text="Count", height=300).update_layout(showlegend=False, margin=dict(t=10)),
            use_container_width=True,
        )

    if not log.empty:
        st.subheader("Recent activity")
        st.dataframe(log.tail(10).iloc[::-1], use_container_width=True, hide_index=True)


def page_contacts():
    st.title("👥 Contacts")
    df = load_master()
    if df.empty:
        st.warning("No data.")
        return

    # Filters
    c1, c2, c3, c4 = st.columns(4)
    tier_filter = c1.multiselect("Tier", sorted(df["Tier"].dropna().unique()), default=[])
    status_options = sorted(df["Status"].replace("", "Not contacted").unique())
    status_filter = c2.multiselect("Status", status_options, default=[])
    min_fit = c3.slider("Min fit_score", -10, 15, value=3)
    search = c4.text_input("Search (Name / Email / URL)")

    filtered = df.copy()
    filtered["Status"] = filtered["Status"].replace("", "Not contacted")
    if tier_filter:
        filtered = filtered[filtered["Tier"].isin(tier_filter)]
    if status_filter:
        filtered = filtered[filtered["Status"].isin(status_filter)]
    filtered = filtered[filtered["fit_score"] >= min_fit]
    if search:
        q = search.lower()
        mask = (
            filtered["Name"].astype(str).str.lower().str.contains(q, na=False)
            | filtered["Email"].astype(str).str.lower().str.contains(q, na=False)
            | filtered["URL"].astype(str).str.lower().str.contains(q, na=False)
        )
        filtered = filtered[mask]

    st.caption(f"Showing {len(filtered)} of {len(df)} contacts")

    # Color-key legend
    st.markdown(
        "<div style='font-size:12px; color:#666;'>"
        "🟡 Sent intro &nbsp;&nbsp; "
        "🟠 Sent program info &nbsp;&nbsp; "
        "🔵 Replied &nbsp;&nbsp; "
        "🔷 Wants money &nbsp;&nbsp; "
        "🟢 Live collab &nbsp;&nbsp; "
        "🔴 Dead &nbsp;&nbsp; "
        "⚪ Not contacted"
        "</div>",
        unsafe_allow_html=True,
    )

    # Display columns (show key info; full table available via toggle)
    show_cols = ["Name", "Email", "Platform", "Tier", "Niche", "Followers", "Status",
                 "First Contact Date", "Last Follow-up", "fit_score", "URL", "Notes"]
    show_cols = [c for c in show_cols if c in filtered.columns]
    if "Status" in df.columns:
        # Row coloring by Status — applied via pandas Styler.
        # Always force dark text on light backgrounds so it's readable in BOTH
        # Streamlit dark and light themes (without this, dark mode = white text
        # on pastel bg = unreadable).
        STATUS_COLORS = {
            "Sent intro":              "background-color: #FFF3CD; color: #1a1a1a",  # soft yellow
            "Sent program info":       "background-color: #FFCC99; color: #1a1a1a",  # orange — escalation
            "Replied":                 "background-color: #D1ECF1; color: #1a1a1a",  # teal/cyan
            "Wants money for posts":   "background-color: #C4DEF2; color: #1a1a1a",  # blue
            "Live collab":             "background-color: #D9F1CF; color: #1a1a1a",  # green
            "Dead":                    "background-color: #F2D4D4; color: #1a1a1a",  # red
            "Do not contact":          "background-color: #F2D4D4; color: #1a1a1a",  # red
        }
        def _color_row(row):
            return [STATUS_COLORS.get(row.get("Status", ""), "")] * len(row)

        styled = filtered[show_cols].style.apply(_color_row, axis=1)

        # Editable status only (other cols read-only)
        edited = st.data_editor(
            styled,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Status": st.column_config.SelectboxColumn(
                    "Status",
                    options=["Not contacted", "Sent intro", "Replied", "Sent program info",
                            "Wants money for posts", "Live collab", "Dead", "Do not contact"],
                    required=True,
                ),
                "URL": st.column_config.LinkColumn("URL"),
                "Followers": st.column_config.NumberColumn(format="%d"),
                "fit_score": st.column_config.NumberColumn(format="%+d"),
            },
            disabled=[c for c in show_cols if c != "Status"],
            num_rows="fixed",
            height=600,
            key="contacts_editor",
        )

        if st.button("💾 Save status changes", type="primary"):
            # DEFENSIVE: count how many rows would have non-empty Status cleared by this save.
            # Catches the Streamlit data_editor + Styler bug where saving without
            # touching anything can mass-revert Status → "Not contacted" → "".
            would_clear = 0
            for idx in edited.index:
                old = (df.loc[idx, "Status"] or "").strip()
                new = edited.loc[idx, "Status"]
                new_normalized = "" if new == "Not contacted" else new
                if old and not new_normalized:
                    would_clear += 1

            if would_clear >= 5:
                st.error(
                    f"⚠️ Save aborted — would clear Status for **{would_clear} contacts** "
                    f"who currently have a non-empty status. Looks like an unintended mass-reset. "
                    f"If this is intentional, change their Status to 'Dead' or 'Do not contact' instead."
                )
            else:
                # Merge changes back into original df
                for idx in edited.index:
                    df.loc[idx, "Status"] = edited.loc[idx, "Status"] if edited.loc[idx, "Status"] != "Not contacted" else ""
                save_master(df)
            with st.spinner("Pushing status to Google Sheet + repainting..."):
                result = subprocess.run(
                    [sys.executable, str(ROOT / "sheet_status_sync.py"), "--paint"],
                    capture_output=True, text=True, cwd=str(ROOT),
                )
            if result.returncode == 0:
                st.success("Saved + synced to Google Sheet.")
            else:
                st.warning("Saved to MASTER, but sheet sync had issues:")
                st.code(result.stderr[-300:] or result.stdout[-300:])
            st.rerun()


def page_send_queue():
    st.title("📬 Send Queue")
    df = load_master()
    cfg = load_config()
    if df.empty:
        st.warning("No data.")
        return

    stage = st.radio("Stage", ["Welcome (cold)", "Offer (after reply)"], horizontal=True)
    min_fit = st.slider("Min fit_score", -10, 15, value=cfg.get("filters", {}).get("min_fit_score", 3))

    if stage == "Welcome (cold)":
        skip = set(cfg.get("filters", {}).get("skip_statuses_welcome", cfg.get("filters", {}).get("skip_statuses", [])))
        queue = df[
            (df["Email"].astype(str).str.strip() != "")
            & (~df["Status"].isin(skip))
            & (df["fit_score"] >= min_fit)
        ].copy()
        sender_arg = ""  # default mode = welcome
        body_tpl = cfg.get("body", "")
        intros = cfg.get("intros", {})
        subject_pool = cfg.get("subjects", [cfg.get("subject", "")])
    else:
        keep = set(cfg.get("filters", {}).get("send_offer_statuses", ["Replied"]))
        queue = df[
            (df["Email"].astype(str).str.strip() != "")
            & (df["Status"].isin(keep))
        ].copy()
        offer = cfg.get("offer", {})
        body_tpl = offer.get("body", "")
        intros = {}
        subject_pool = offer.get("subjects", [])

    queue = queue.sort_values(["Tier", "fit_score", "Followers"], ascending=[True, False, False])

    st.caption(f"Queue size: **{len(queue)}** | Daily limit: {cfg.get('sender', {}).get('daily_limit', 80)}")

    if len(queue) == 0:
        st.info("Queue is empty.")
        return

    # Preview first N
    n_preview = st.number_input("Preview / send batch size", 1, 50, value=5)
    preview = queue.head(n_preview)
    st.dataframe(
        preview[["Name", "Email", "Tier", "Followers", "fit_score", "Status"]],
        use_container_width=True, hide_index=True,
    )

    # Show subject + intro for tier of first contact
    if not preview.empty and intros:
        st.subheader("Subject rotation preview")
        for i in range(min(len(preview), len(subject_pool))):
            row = preview.iloc[i]
            subj = subject_pool[i % len(subject_pool)].replace("{Name}", str(row["Name"]))
            st.code(f"#{i+1}  {row['Name']}  ({row['Tier']})\nSubject: {subj}", language=None)

    st.divider()
    st.markdown("### Send batch")
    st.warning(
        "⚠️ This will **actually send emails**. Make sure your templates and queue are correct. "
        "Status will update to 'Sent intro' after each send."
    )
    confirm_text = st.text_input(f"Type **SEND** to confirm sending {n_preview} emails")
    if st.button("🚀 Send batch", type="primary", disabled=(confirm_text != "SEND")):
        with st.spinner(f"Sending {n_preview} emails..."):
            cmd = [sys.executable, str(ROOT / "outreach_sender.py"), "--send",
                   "--limit", str(n_preview), "--source", str(MASTER_FILE)]
            cmd += ["--stage", "offer" if stage.startswith("Offer") else "welcome"]
            result = subprocess.run(cmd, capture_output=True, text=True, cwd=str(ROOT))
        st.code(result.stdout + ("\n" + result.stderr if result.stderr else ""))
        if result.returncode == 0:
            st.success("Done. Check log and contacts page.")
            st.cache_data.clear()
        else:
            st.error(f"Sender exited with code {result.returncode}")


def page_templates():
    st.title("📝 Templates")
    cfg = load_config()
    if not cfg:
        st.error("outreach_config.yaml not found.")
        return

    tabs = st.tabs(["Welcome (cold)", "Offer (after reply)", "Raw YAML"])

    with tabs[0]:
        st.subheader("Subject variations (rotated round-robin)")
        for i, s in enumerate(cfg.get("subjects", []), 1):
            st.code(f"#{i}  {s}", language=None)

        st.subheader("Body template")
        st.code(cfg.get("body", ""), language=None)

        st.subheader("Tier-specific intro sentences")
        for tier, intro in (cfg.get("intros") or {}).items():
            st.markdown(f"**Tier {tier}**")
            st.code(intro, language=None)

    with tabs[1]:
        offer = cfg.get("offer", {})
        st.subheader("Subjects")
        for i, s in enumerate(offer.get("subjects", []), 1):
            st.code(f"#{i}  {s}", language=None)
        st.subheader("Body")
        st.code(offer.get("body", ""), language=None)
        st.info("Offer template currently has TODO placeholders — fill in real program details.")

    with tabs[2]:
        st.code(CONFIG_FILE.read_text(encoding="utf-8"), language="yaml")
        st.caption(f"Edit `{CONFIG_FILE.name}` directly to modify templates.")


def page_log():
    st.title("📜 Send Log")
    log = load_log()
    if log.empty:
        st.info("No sends yet.")
        return

    c1, c2 = st.columns(2)
    outcomes = sorted(log["outcome"].unique())
    outcome_filter = c1.multiselect("Outcome", outcomes, default=outcomes)
    tier_filter = c2.multiselect("Tier", sorted(log["tier"].astype(str).unique()), default=[])

    filtered = log[log["outcome"].isin(outcome_filter)]
    if tier_filter:
        filtered = filtered[filtered["tier"].astype(str).isin(tier_filter)]

    st.caption(f"Showing {len(filtered)} of {len(log)} entries")
    st.dataframe(filtered.iloc[::-1], use_container_width=True, hide_index=True)

    st.download_button(
        "⬇ Download log as CSV",
        data=filtered.to_csv(index=False).encode("utf-8"),
        file_name=f"outreach_log_{datetime.now().strftime('%Y%m%d')}.csv",
        mime="text/csv",
    )


# ----------------------------------------------------------------------
# Main router
# ----------------------------------------------------------------------

st.set_page_config(page_title="Timestripe Outreach Admin", page_icon="🎯", layout="wide")

st.sidebar.title("🎯 Timestripe Outreach")
st.sidebar.caption(f"Master file: `{MASTER_FILE.name}`")

# Single "Refresh & Sync" button — pulls new replies from IMAP, applies them
# to MASTER, syncs to Google Sheet, repaints colors, and refreshes the UI.
# This is what users intuitively expect "Refresh" to do.
if st.sidebar.button("🔄 Refresh & Sync", help="Pull new replies from Yandex, sync to Google Sheet, refresh UI"):
    with st.spinner("Checking inbox for replies + syncing..."):
        result = subprocess.run(
            [sys.executable, str(ROOT / "reply_detector.py"), "--apply", "--since", "14"],
            capture_output=True, text=True, cwd=str(ROOT), timeout=180,
        )
    if result.returncode == 0:
        st.sidebar.success("✓ Refreshed")
    else:
        st.sidebar.warning("Sync had issues — see below")
    # Show last bit of output (mostly: how many replies/sent found)
    out = result.stdout or result.stderr
    if out:
        st.sidebar.code(out[-700:])
    st.cache_data.clear()
    st.rerun()

# Lightweight UI-only refresh (no IMAP poll) — for when you've just edited
# Status in Contacts and want to re-render without re-syncing
if st.sidebar.button("👁 Reload UI only", help="Re-read MASTER from disk; does NOT poll mailbox"):
    st.cache_data.clear()
    st.rerun()

if st.sidebar.button("🔍 Run YouTube parser (delta)"):
    with st.spinner("Running YouTube parser..."):
        result = subprocess.run(
            [sys.executable, str(ROOT / "youtube_finder.py")],
            capture_output=True, text=True, cwd=str(ROOT),
        )
    st.sidebar.code(result.stdout[-800:] if result.stdout else result.stderr[-800:])

if st.sidebar.button("🎨 Paint Google Sheet"):
    with st.spinner("Painting rows by Status..."):
        result = subprocess.run(
            [sys.executable, str(ROOT / "paint_sheet.py")],
            capture_output=True, text=True, cwd=str(ROOT),
        )
    st.sidebar.code(result.stdout[-500:] if result.stdout else result.stderr[-500:])

PAGE = st.sidebar.radio(
    "Page",
    ["📊 Overview", "👥 Contacts", "📬 Send Queue", "📝 Templates", "📜 Send Log"],
)

st.sidebar.divider()
st.sidebar.caption(
    "**CLI shortcuts**\n\n"
    "```bash\n"
    "# Re-parse YouTube\n"
    "python youtube_finder.py\n\n"
    "# Send N emails\n"
    "python outreach_sender.py --send --limit 5\n"
    "```"
)

if PAGE == "📊 Overview":
    page_overview()
elif PAGE == "👥 Contacts":
    page_contacts()
elif PAGE == "📬 Send Queue":
    page_send_queue()
elif PAGE == "📝 Templates":
    page_templates()
elif PAGE == "📜 Send Log":
    page_log()
